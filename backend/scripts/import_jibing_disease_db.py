"""
疾病数据库.xlsx 导入脚本
从 Excel 文件导入疾病数据到数据库
"""
import sys
import os
from pathlib import Path

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from sqlalchemy.orm import Session
from openpyxl import load_workbook
from pypinyin import lazy_pinyin, Style
import re

from app.database import get_db
from app.models.disease import Disease
from app.models.department import Department


# 科室映射表 - 将 Excel 中的科室名称映射到标准科室
DEPARTMENT_MAPPING = {
    # 内科
    "神经内科": "神经内科",
    "精神病科": "精神心理科",
    "精神心理科": "精神心理科",
    "呼吸内科": "呼吸内科",
    "心血管内科": "心血管内科",
    "消化内科": "消化内科",
    "肾内科": "肾内科",
    "内分泌科": "内分泌科",
    "血液科": "血液科",
    "风湿科": "风湿科",
    "传染科": "感染科",
    "感染科": "感染科",
    "变态反应科": "变态反应科",

    # 外科
    "骨科": "骨科",
    "普外科": "普外科",
    "心胸外科": "心胸外科",
    "神经外科": "神经外科",
    "泌尿外科": "泌尿外科",
    "肛肠外科": "肛肠外科",
    "外伤科": "急诊外科",
    "急诊外科": "急诊外科",
    "烧伤科": "烧伤科",
    "整形外科": "整形外科",
    "显微外科": "显微外科",
    "手外科": "手外科",

    # 妇产科
    "妇科": "妇科",
    "妇产科": "妇产科",
    "产科": "产科",
    "不孕不育": "生殖医学科",
    "生殖医学科": "生殖医学科",

    # 儿科
    "小儿科": "儿科",
    "儿科": "儿科",
    "新生儿科": "新生儿科",

    # 五官科
    "眼科": "眼科",
    "耳鼻喉": "耳鼻喉科",
    "耳鼻喉科": "耳鼻喉科",
    "口腔科": "口腔科",
    "五官科": "五官科",

    # 皮肤性病科
    "皮肤科": "皮肤科",
    "性病科": "性病科",

    # 肿瘤科
    "肿瘤科": "肿瘤科",
    "放疗、化疗科": "肿瘤科",
    "放疗科": "肿瘤科",
    "化疗科": "肿瘤科",

    # 中医科
    "中医科": "中医科",
    "中西医结合科": "中西医结合科",

    # 其他
    "急诊科": "急诊科",
    "重症监护室": "重症医学科",
    "重症医学科": "重症医学科",
    "康复科": "康复科",
    "介入科": "介入科",
    "职业病科": "职业病科",
    "心理咨询": "心理门诊",
    "心理门诊": "心理门诊",
    "男科": "男科",
    "成瘾医学科": "成瘾医学科",
    "疼痛科": "疼痛科",
    "老年病科": "老年病科",
    "全科医学科": "全科医学科",
    "预防保健科": "预防保健科",

    # 影像/检验（不算临床科室，但保留）
    "CT室": "放射科",
    "MR室": "放射科",
    "X光室": "放射科",
    "放射科": "放射科",
    "检验科": "检验科",
}


def get_standard_department(raw_dept: str, db: Session) -> Department:
    """
    从原始科室字符串获取标准科室
    处理逻辑：
    1. 如果是单个科室，直接映射
    2. 如果是多个科室（用空格/制表符分隔），取第一个有效的
    3. 如果数据库中没有，创建新科室
    """
    if not raw_dept:
        return None

    # 分割科室（按空格、制表符、全角空格）
    parts = re.split(r'[\s　]+', raw_dept.strip())
    parts = [p for p in parts if p]  # 去除空字符串

    for part in parts:
        # 查找映射
        standard_name = DEPARTMENT_MAPPING.get(part, part)

        # 在数据库中查找
        dept = db.query(Department).filter(Department.name == standard_name).first()

        if dept:
            return dept

        # 如果没有找到，创建新科室
        new_dept = Department(
            name=standard_name,
            description=f"从疾病数据库导入",
            sort_order=99,
            is_primary=False
        )
        db.add(new_dept)
        db.commit()
        db.refresh(new_dept)
        print(f"  [新建科室] {standard_name}")
        return new_dept

    return None


def generate_pinyin(text: str) -> tuple:
    """生成拼音和拼音首字母"""
    if not text:
        return None, None

    # 全拼音
    pinyin_list = lazy_pinyin(text, style=Style.NORMAL)
    pinyin_full = ''.join(pinyin_list)

    # 首字母
    pinyin_abbr_list = lazy_pinyin(text, style=Style.FIRST_LETTER)
    pinyin_abbr = ''.join(pinyin_abbr_list)

    return pinyin_full, pinyin_abbr


def clean_content(text: str) -> str:
    """清理内容，去除多余的标题前缀"""
    if not text:
        return None

    # 去除类似 "阿尔采末病症状_阿尔采末病有什么症状\n\n" 这样的前缀
    text = re.sub(r'^[^\n_]+_[^\n]+\n\n', '', text)

    # 去除多余的空白行
    text = re.sub(r'\n{3,}', '\n\n', text)

    # 去除首尾空白
    text = text.strip()

    return text if text else None


def import_diseases_from_excel(excel_path: str, db: Session, dry_run: bool = False):
    """从 Excel 导入疾病数据"""

    print(f"正在读取 Excel 文件: {excel_path}")
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    total_rows = ws.max_row - 1  # 减去表头
    print(f"总数据行数: {total_rows}")
    print()

    imported = 0
    skipped = 0
    errors = []

    # 获取所有现有科室
    existing_departments = {d.name: d for d in db.query(Department).all()}
    print(f"现有科室数: {len(existing_departments)}")
    print()

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:  # 跳过表头
            continue

        # 读取数据
        # 列顺序: 疾病名称, 别名, 简介, 病因, 症状及病史, 诊断, 并发症, 治疗, 预防, 部位, 科室, 相关症状, 检查项目, 相关疾病
        name = row[0]
        aliases = row[1]
        overview = row[2]
        causes = row[3]
        symptoms = row[4]
        diagnosis = row[5]
        complications = row[6]
        treatment = row[7]
        prevention = row[8]
        body_part = row[9]
        department_raw = row[10]
        related_symptoms = row[11]
        exam_items = row[12]
        related_diseases = row[13]

        if not name:
            skipped += 1
            continue

        # 检查是否已存在
        existing = db.query(Disease).filter(Disease.name == name).first()
        if existing:
            skipped += 1
            if skipped <= 10:
                print(f"  [跳过] 已存在: {name}")
            continue

        # 处理科室
        department = get_standard_department(department_raw, db)

        # 生成拼音
        pinyin_full, pinyin_abbr = generate_pinyin(name)

        # 清理内容
        overview_clean = clean_content(overview)
        causes_clean = clean_content(causes)
        symptoms_clean = clean_content(symptoms)
        diagnosis_clean = clean_content(diagnosis)
        treatment_clean = clean_content(treatment)
        prevention_clean = clean_content(prevention)

        # 清理并发症（单独字段）
        complications_clean = clean_content(complications)

        # 清理其他扩展字段
        body_parts_clean = str(body_part).strip() if body_part else None
        related_symptoms_clean = str(related_symptoms).strip() if related_symptoms else None
        exam_items_clean = str(exam_items).strip() if exam_items else None
        related_diseases_clean = str(related_diseases).strip() if related_diseases else None

        # 创建疾病记录
        disease = Disease(
            name=name,
            aliases=aliases,
            pinyin=pinyin_full,
            pinyin_abbr=pinyin_abbr,
            department_id=department.id if department else None,
            recommended_department=department_raw,
            overview=overview_clean,
            symptoms=symptoms_clean,
            causes=causes_clean,
            diagnosis=diagnosis_clean,
            treatment=treatment_clean,
            prevention=prevention_clean,
            care=care_clean,
            complications=complications_clean,
            body_parts=body_parts_clean,
            related_symptoms=related_symptoms_clean,
            exam_items=exam_items_clean,
            related_diseases=related_diseases_clean,
            source="jibing_disease_db",
            is_hot=False,
            is_active=True,
            sort_order=0,
            view_count=0
        )

        if not dry_run:
            try:
                db.add(disease)
                db.commit()
                imported += 1
                if imported <= 10 or imported % 100 == 0:
                    dept_name = department.name if department else department_raw
                    print(f"  [{imported}] {name} -> {dept_name}")
            except Exception as e:
                db.rollback()
                errors.append(f"{name}: {str(e)}")
                print(f"  [错误] {name}: {e}")
        else:
            imported += 1

    # 打印统计
    print()
    print("=" * 50)
    print("导入完成!")
    print(f"  导入成功: {imported}")
    print(f"  跳过（已存在）: {skipped}")
    if errors:
        print(f"  错误: {len(errors)}")
        for err in errors[:10]:
            print(f"    - {err}")
    print("=" * 50)


def main():
    import argparse

    parser = argparse.ArgumentParser(description="导入疾病数据库.xlsx")
    parser.add_argument("--dry-run", action="store_true", help="试运行，不实际写入数据库")
    parser.add_argument("--file", default="data/疾病数据库.xlsx", help="Excel 文件路径")
    args = parser.parse_args()

    # 获取数据库连接
    db_gen = get_db()
    db = next(db_gen)

    try:
        import_diseases_from_excel(args.file, db, dry_run=args.dry_run)
    finally:
        db.close()


if __name__ == "__main__":
    main()

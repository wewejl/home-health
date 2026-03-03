#!/usr/bin/env python3
"""
药品数据库导入脚本
从 Excel 文件导入药品数据到数据库
"""
import os
import sys

# 添加项目路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from sqlalchemy.orm import Session
from app.database import engine, Base, SessionLocal
from app.models import Drug, DrugCategory


def get_or_create_category(db: Session, name: str) -> DrugCategory:
    """获取或创建分类"""
    category = db.query(DrugCategory).filter(DrugCategory.name == name).first()
    if not category:
        category = DrugCategory(
            name=name,
            icon=None,
            description=f"{name}类药品",
            display_type="grid",
            sort_order=0,
            is_active=True
        )
        db.add(category)
        db.commit()
        db.refresh(category)
        print(f"  [分类] 创建: {name}")
    return category


def clean_value(value):
    """清理单元格值"""
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() if value.strip() else None
    return str(value) if value not in [None, ""] else None


def import_drugs_from_excel(excel_path: str, batch_size: int = 500):
    """
    从 Excel 导入药品数据

    Args:
        excel_path: Excel 文件路径
        batch_size: 每批次提交的记录数
    """
    print(f"正在读取 Excel 文件: {excel_path}")
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    print(f"总行数: {ws.max_row}")
    print("开始导入...\n")

    db = SessionLocal()

    try:
        # 首先收集所有用途分类
        print("=== 步骤1: 收集用途分类 ===")
        usage_categories = set()
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            # Excel 列索引 (0-based): 10 = 用途分类
            usage = clean_value(row[10])
            if usage and usage != "未分类":
                usage_categories.add(usage)

        print(f"发现 {len(usage_categories)} 个用途分类")
        for cat in sorted(usage_categories):
            get_or_create_category(db, cat)

        print("\n=== 步骤2: 导入药品数据 ===")

        # Excel 列索引映射 (0-based)
        COL = {
            '序号': 0,
            '商品条码': 1,
            '药品名称': 2,
            '拼音简码': 3,
            '规格': 4,
            '剂型': 5,
            '包装单位': 6,
            '批准文号': 7,
            '处方类型': 8,
            '性质分类': 9,
            '用途分类': 10,
            '商品名商标': 11,
            '主要成分': 12,
            '性状': 13,
            '适应症': 14,
            '用法用量': 15,
            '不良反应': 16,
            '禁忌': 17,
            '注意事项': 18,
            '药物相互作用': 19,
            '贮藏': 20,
            '生产厂家': 21,
            '产地': 22,
            '本位码': 23,
        }

        imported_count = 0
        skipped_count = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 1):
            # 获取药品名称（必需字段）
            drug_name = clean_value(row[COL['药品名称']])
            if not drug_name:
                skipped_count += 1
                continue

            # 处理用途分类关联
            usage_category = clean_value(row[COL['用途分类']])
            categories_list = []
            if usage_category and usage_category != "未分类":
                category = get_or_create_category(db, usage_category)
                categories_list.append(category)

            # 创建药品对象
            drug = Drug(
                name=drug_name,
                pinyin_abbr=clean_value(row[COL['拼音简码']]),
                common_brands=clean_value(row[COL['商品名商标']]),
                # 新增字段
                barcode=clean_value(row[COL['商品条码']]),
                approval_number=clean_value(row[COL['批准文号']]),
                specification=clean_value(row[COL['规格']]),
                dosage_form=clean_value(row[COL['剂型']]),
                package_unit=clean_value(row[COL['包装单位']]),
                prescription_type=clean_value(row[COL['处方类型']]),
                drug_nature=clean_value(row[COL['性质分类']]),
                ingredients=clean_value(row[COL['主要成分']]),
                appearance=clean_value(row[COL['性状']]),
                manufacturer=clean_value(row[COL['生产厂家']]),
                origin=clean_value(row[COL['产地']]),
                standard_code=clean_value(row[COL['本位码']]),
                # 内容模块
                indications=clean_value(row[COL['适应症']]),
                dosage=clean_value(row[COL['用法用量']]),
                side_effects=clean_value(row[COL['不良反应']]),
                contraindications=clean_value(row[COL['禁忌']]),
                precautions=clean_value(row[COL['注意事项']]),
                interactions=clean_value(row[COL['药物相互作用']]),
                storage=clean_value(row[COL['贮藏']]),
                # 默认值
                is_active=True,
                is_hot=False,
                sort_order=0,
                view_count=0,
            )

            # 添加分类关联
            for cat in categories_list:
                drug.categories.append(cat)

            # 单个保存（确保关系被保存）
            db.add(drug)
            db.commit()
            db.refresh(drug)

            imported_count += 1

            # 进度显示
            if imported_count % 1000 == 0:
                print(f"  进度: {row_idx}/{ws.max_row - 1} (已导入: {imported_count})")

        print(f"\n=== 导入完成 ===")
        print(f"总处理: {ws.max_row - 1} 条")
        print(f"成功导入: {imported_count} 条")
        print(f"跳过已存在: {skipped_count} 条")

        # 统计数据库中的药品数量
        total_drugs = db.query(Drug).count()
        total_categories = db.query(DrugCategory).count()
        print(f"\n数据库统计:")
        print(f"  药品总数: {total_drugs}")
        print(f"  分类总数: {total_categories}")

    except Exception as e:
        db.rollback()
        print(f"\n错误: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        db.close()
        wb.close()


if __name__ == "__main__":
    excel_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), "药品库带条码.xlsx")

    if not os.path.exists(excel_file):
        print(f"错误: 文件不存在: {excel_file}")
        sys.exit(1)

    print("=" * 50)
    print("药品数据库导入工具")
    print("=" * 50)

    # 创建数据库表（如果不存在）
    Base.metadata.create_all(bind=engine)

    # 开始导入
    import_drugs_from_excel(excel_file)

    print("\n导入完成!")

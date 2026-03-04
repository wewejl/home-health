"""
疾病数据库导入脚本
从 疾病数据库.xlsx 导入数据到数据库
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from tqdm import tqdm
from app.database import SessionLocal, engine, Base
from app.models.disease import Disease
from app.models.department import Department
from sqlalchemy import text


def get_department_id(db: SessionLocal, dept_name: str) -> int | None:
    """根据科室名称获取科室ID（支持多科室匹配）"""
    if not dept_name:
        return None
    # 清理名称，去除空格和特殊字符
    clean_name = dept_name.strip().replace('　', ' ').replace('\t', ' ')
    if not clean_name:
        return None

    # 尝试精确匹配
    dept = db.query(Department).filter(Department.name == clean_name).first()
    if dept:
        return dept.id

    # 尝试从多科室字符串中提取匹配的科室
    # 按空格分割科室名称
    parts = clean_name.split()
    all_depts = db.query(Department).all()
    dept_names = {d.name.lower(): d.id for d in all_depts}

    for part in parts:
        if part.lower() in dept_names:
            return dept_names[part.lower()]

    # 尝试模糊匹配（包含）
    for dept in all_depts:
        if dept.name.lower() in clean_name.lower():
            return dept.id

    return None


def import_diseases_from_excel(file_path: str, batch_size: int = 100):
    """从Excel导入疾病数据"""
    db = SessionLocal()

    try:
        print(f"正在加载文件: {file_path}")
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active

        total_rows = ws.max_row - 1  # 减去标题行
        print(f"总数据行数: {total_rows}")

        # 先获取所有科室映射
        print("\n正在构建科室映射...")
        dept_cache = {}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            dept_name = row[10]  # 科室列
            if dept_name and dept_name.strip() and dept_name.strip() not in dept_cache:
                dept_id = get_department_id(db, dept_name.strip())
                dept_cache[dept_name.strip()] = dept_id
                if dept_id:
                    print(f"  ✓ {dept_name.strip()} -> ID {dept_id}")
                else:
                    print(f"  ✗ {dept_name.strip()} -> 未找到")

        # 清空现有数据（可选）
        print("\n清空现有的 disease_db 来源数据...")
        result = db.execute(text("DELETE FROM diseases WHERE source = :source"), {"source": "disease_db"})
        db.commit()
        print(f"已清空 {result.rowcount} 条旧数据")

        # 批量导入
        diseases_to_create = []
        skipped = 0

        print("\n开始导入...")
        for row in tqdm(list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))):
            # 解析行数据
            (name, aliases, overview, causes, symptoms, diagnosis,
             complications, treatment, prevention, body_parts,
             dept_name, related_symptoms, exam_items, related_diseases) = row

            if not name or not name.strip():
                skipped += 1
                continue

            disease_data = {
                "name": name.strip(),
                "aliases": aliases.strip() if aliases else None,
                "overview": overview.strip() if overview else None,
                "causes": causes.strip() if causes else None,
                "symptoms": symptoms.strip() if symptoms else None,
                "diagnosis": diagnosis.strip() if diagnosis else None,
                "complications": complications.strip() if complications else None,
                "treatment": treatment.strip() if treatment else None,
                "prevention": prevention.strip() if prevention else None,
                "body_parts": body_parts.strip() if body_parts else None,
                "recommended_department": dept_name.strip() if dept_name else None,
                "related_symptoms": related_symptoms.strip() if related_symptoms else None,
                "exam_items": exam_items.strip() if exam_items else None,
                "related_diseases": related_diseases.strip() if related_diseases else None,
                "source": "disease_db",
                "is_active": True,
            }

            # 查找科室ID
            if dept_name and dept_name.strip():
                disease_data["department_id"] = dept_cache.get(dept_name.strip())

            diseases_to_create.append(Disease(**disease_data))

            # 批量保存
            if len(diseases_to_create) >= batch_size:
                db.bulk_save_objects(diseases_to_create)
                db.commit()
                diseases_to_create.clear()

        # 保存剩余数据
        if diseases_to_create:
            db.bulk_save_objects(diseases_to_create)
            db.commit()

        wb.close()

        # 统计结果
        total = db.query(Disease).filter(Disease.source == "disease_db").count()
        print(f"\n✅ 导入完成!")
        print(f"  - 导入总数: {total}")
        print(f"  - 跳过空名: {skipped}")

    except Exception as e:
        db.rollback()
        print(f"\n❌ 导入失败: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    file_path = "/app/data/疾病数据库.xlsx"

    if not os.path.exists(file_path):
        print(f"❌ 文件不存在: {file_path}")
        sys.exit(1)

    import_diseases_from_excel(file_path)
